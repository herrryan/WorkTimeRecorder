import sys
import time
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Style, PatternFill, Alignment
from openpyxl.styles.colors import RED

from TimeEntryTool import * 
from PyQt4 import QtCore, QtGui

reload(sys)
sys.setdefaultencoding('utf8')

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

headers = ['Date', 'Time (h)', 'Task', 'for whom? (mei62, Housing office, others)'];
entry_list = [];
class Entry:
    __entryCount = 0;
    def __init__(self, date, time, task, for_whom):
        self.date = date;
        self.time = time;
        self.task = task;
        self.for_whom = for_whom;
        self.__entryCount += 1;

class MainDialog(QtGui.QDialog, Ui_baseWindow):
    date = 'dasd';
    hours = 0.0;
    task = '';
    for_whom = '';
    def __init__(self, parent=None):
        super(MainDialog, self).__init__(parent);
        self.setupUi(self);
        # default
        self.dateEdit.setDate(QtCore.QDate.currentDate());
        self.dateEdit.setMinimumDate(QtCore.QDate.currentDate().addDays(-365));
        self.dateEdit.setMaximumDate(QtCore.QDate.currentDate().addDays(365));
        self.dateEdit.setDisplayFormat('dd.MM.yyyy');

        # controls
        self.row = 1;
        self.pushButton.clicked.connect(self.add_new_line_and_append_to_entry_list, self.row);
        self.get_date();
        self.dateEdit.dateChanged.connect(self.get_date);
        self.get_hours();
        self.doubleSpinBox.valueChanged.connect(self.get_hours);
        self.get_task();
        self.lineEdit.textChanged.connect(self.get_task);
        self.get_for_whom();
        self.comboBox.currentIndexChanged.connect(self.get_for_whom);
        #TODO: create entry and further create entry list for printing
        #create_time_entry(date, time, task, for_whom);
    def add_new_line_and_append_to_entry_list(self, row):
        self.row += 1;
        self.add_new_line(self.row);
        entry = [self.date, self.hours, self.task, self.for_whom];
        entry_list.append(entry);
        self.setup_workbook("stundenliste china-tutorin1", "July", entry_list);
        print "entry %s" % entry;
        print "add_new_line_and_append_to_entry_list";
    def get_date(self):
        self.date = str(self.dateEdit.date().toString());
        print "date %s" % self.dateEdit.date().toString();
    def get_hours(self):
        self.hours = self.doubleSpinBox.value();
        print "hours %.1f" % self.doubleSpinBox.value();
    def get_task(self):
        self.task = str(self.lineEdit.text());
        print "task %s" % self.lineEdit.text();
    def get_for_whom(self):
        self.for_whom = str(self.comboBox.currentText());
        print "from %s" % self.comboBox.currentText();
    
    def add_new_line(self, row):
        print "row number %s" % row
        # TODO handle multiple with a list
        self.comboBox = QtGui.QComboBox(self.gridLayoutWidget_2)
        self.comboBox.setObjectName(_fromUtf8("comboBox"))
        self.comboBox.addItem(_fromUtf8(""))
        self.comboBox.addItem(_fromUtf8(""))
        self.comboBox.addItem(_fromUtf8(""))
        # the four number is [row, relative_position, component_height, component_width]
        self.gridLayout.addWidget(self.comboBox, row, 3, 1, 1)
        self.dateEdit = QtGui.QDateEdit(self.gridLayoutWidget_2)
        self.dateEdit.setCalendarPopup(True)
        self.dateEdit.setObjectName(_fromUtf8("dateEdit"))
        self.gridLayout.addWidget(self.dateEdit, row, 0, 1, 1)
        self.lineEdit = QtGui.QLineEdit(self.gridLayoutWidget_2)
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.lineEdit.setPlaceholderText("what have you done?")
        self.gridLayout.addWidget(self.lineEdit, row, 2, 1, 1)
        self.doubleSpinBox = QtGui.QDoubleSpinBox(self.gridLayoutWidget_2)
        self.doubleSpinBox.setObjectName(_fromUtf8("doubleSpinBox"))
        self.gridLayout.addWidget(self.doubleSpinBox, row, 1, 1, 1)

    def setup_default_layout(self, worksheet, entry):
        worksheet.cell('A3').value = 'WOKO Time recording';
        worksheet.cell('A3').font = Font(bold=True);
        for prefix in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            worksheet.cell(prefix + '3').fill = PatternFill(fill_type='solid', start_color='C0C0C0');
        worksheet.cell('A6').value = 'Student liaison assistant';
        worksheet.cell('A6').font = Font(bold=True);
        worksheet.cell('A7').value = 'Zhiyang Yu from ____ to _____';
        worksheet.cell('A7').font = Font(bold=True);
        worksheet.append(headers);
        for entry in entry_list:
            worksheet.append(entry);
        for prefix in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            worksheet.cell(prefix + '20').border = Border(bottom=Side(border_style='thin'));
        worksheet.cell('A22').value = 'Total';
        worksheet.cell('A22').font = Font(bold=True);
        worksheet.cell('B22').value = 0.0;
        worksheet.cell('B22').font = Font(bold=True);
        worksheet.cell('B22').border = Border(bottom=Side(border_style='double'));
        worksheet.cell('B22').alignment = Alignment(horizontal='left');
        worksheet.cell('C22').value = 'hours';
        worksheet.cell('C22').border = Border(bottom=Side(border_style='double'));
        worksheet.cell('E22').value = 'Visa Head of Operations: ____________';
        worksheet.cell('A25').value = 'To be sent to the WOKO office until the 20th of each month (e-mail).';
        worksheet.cell('A25').font = Font(color=RED);

    def setup_workbook(self, file_name, sheet_name, entry_list):
        workbook = Workbook();
        worksheet = workbook.worksheets[0];
        worksheet.title = sheet_name;
        total_hours = 0.0;
        self.setup_default_layout(worksheet, entry_list);
        save_path = file_name + '.xlsx';
        workbook.save(save_path);

def create_time_entry(date, time, task, for_whom):
    return Entry(date, time, task, for_whom);

def export_to_woko(entry_list):
    for entry in entry_list:
        return 1;
def print_line_in_workbook(entry):
    return 1;



if __name__=='__main__':
    app = QtGui.QApplication(sys.argv)
    window = MainDialog();
    window.show();
    sys.exit(app.exec_())